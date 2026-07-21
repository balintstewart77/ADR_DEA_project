"""Build the restricted wide-format pilot case-review workbook.

The builder is deliberately offline. It joins existing human and model outputs,
retains stored rationale and comment text, and writes data-bearing products only
below ``preregistration_restricted``.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import subprocess
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .schema import DOMAIN_LABELS, PURPOSE_LABELS, UNCLEAR
from .scratch_agreement import complete_set_pattern


REPOSITORY_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_READY = Path(
    "preregistration/package/05_training_and_pilot/pilot_analysis/"
    "pilot_analysis_ready.csv"
)
DEFAULT_AGREEMENT = Path(
    "preregistration/package/05_training_and_pilot/pilot_analysis/"
    "pilot_record_agreement.csv"
)
DEFAULT_DIMENSION_SUMMARY = Path(
    "preregistration/package/05_training_and_pilot/pilot_analysis/"
    "pilot_record_dimension_summary.csv"
)
DEFAULT_RAW = Path(
    "preregistration/package/05_training_and_pilot/"
    "pilot_raw_DATA_2026-07-20_2153.csv"
)
DEFAULT_FABLE = Path(
    "analysis/outputs_classified_20260702_fable5/layer_classifications.csv"
)
DEFAULT_FABLE_METADATA = Path(
    "analysis/outputs_classified_20260702_fable5/run_metadata.json"
)
DEFAULT_GPT_RUN_SCRIPT = Path("analysis/outputs/gpt55_crossmodel_stratum_run.py")
DEFAULT_GPT_REFERENCE = Path(
    "preregistration_restricted/pilot_private_review/"
    "pilot_human_model_classifications.csv"
)
DEFAULT_OUTPUT_DIR = Path("preregistration_restricted/pilot_private_review")

CODERS = ("C01", "C02", "C03")
CLASSIFICATION_DIMENSIONS = ("Research Domains", "Analytical Purposes")
DIAGNOSTIC_DIMENSIONS = ("Confidence", "Register sufficiency", "Taxonomy fit")
EXPECTED_TAXONOMY = "dict-1.0-rc2"
PROVISIONAL_FILENAME = "classifications_1309_precollapse_PROVISIONAL.csv"
PROVISIONAL_WARNING_TITLE = "GPT source status: TEMPORARY PRE-CANONICAL SNAPSHOT"
PROVISIONAL_WARNING = (
    "This restricted workbook was generated using "
    "preregistration_restricted/classifications_1309_precollapse_PROVISIONAL.csv, "
    "an archived 1,309-row pre-collapse GPT-5.5 snapshot.\n\n"
    "It is suitable only for exploratory pilot review. The workbook must be "
    "regenerated and checked against the recovered canonical 1,308-row GPT-5.5 "
    "snapshot before pilot closure, final package freeze, preregistration, "
    "publication or formal reporting."
)
MISSING_RATIONALE = "Not available in stored model output"
KNOWN_WARNING = (
    "Known instrument warning: candidate-0.3 incoherent None taxonomy-issue "
    "selection retained for C02 on 2021/103; not recoded."
)

DOMAIN_ORDER = (
    "Labour Market & Employment",
    "Education & Skills",
    "Health & Social Care",
    "Crime & Justice",
    "Business & Productivity",
    "Poverty, Wealth & Living Standards",
    "Housing & Planning",
    "Migration & Demographics",
    "Environment & Agriculture",
    "Public Finance & Taxation",
    "Data Infrastructure & Methodology",
    UNCLEAR,
)
PURPOSE_ORDER = (
    "Descriptive Monitoring",
    "Outcome Tracking",
    "Life-Course / Trajectory Analysis",
    "Service Interaction / Systems Analysis",
    "Policy Evaluation / Impact Analysis",
    "Risk Prediction / Early Identification",
    "Methodological / Infrastructure Research",
    UNCLEAR,
)

COLORS = {
    "green": "FF5AAE61",
    "red": "FFD6604D",
    "amber": "FFF6C85F",
    "neutral": "FFF2F2F2",
    "warning": "FFFFD966",
    "group": "FF415A77",
    "header": "FFD9E2F3",
    "white": "FFFFFFFF",
}


class CaseReviewError(RuntimeError):
    """Raised when inputs, joins, or restricted-output rules fail."""


@dataclass(frozen=True)
class ColumnSpec:
    group: str
    key: str
    label: str
    width: float


@dataclass(frozen=True)
class ModelSource:
    path: Path
    sha256: str
    byte_size: int
    row_count: int
    unique_record_count: int
    encoding: str
    rows_by_id: Mapping[str, Mapping[str, str]]
    rationale_field: str | None


COLUMNS = (
    ColumnSpec("Record context", "record_id", "Record ID", 15),
    ColumnSpec("Record context", "project_title", "Project title", 38),
    ColumnSpec("Record context", "public_register_text", "Public register text supplied to classification", 58),
    ColumnSpec("Research Domains", "domain_agreement_pattern", "Human agreement pattern", 20),
    ColumnSpec("Research Domains", "domain_c01", "Domain C01", 28),
    ColumnSpec("Research Domains", "domain_c02", "Domain C02", 28),
    ColumnSpec("Research Domains", "domain_c03", "Domain C03", 28),
    ColumnSpec("Research Domains", "domain_fable", "Domain Fable", 28),
    ColumnSpec("Research Domains", "domain_fable_rationale", "Domain Fable stored rationale", 52),
    ColumnSpec("Research Domains", "domain_gpt55", "Domain provisional GPT-5.5", 28),
    ColumnSpec("Research Domains", "domain_gpt55_rationale", "Domain provisional GPT-5.5 stored rationale", 52),
    ColumnSpec("Analytical Purposes", "purpose_agreement_pattern", "Human agreement pattern", 20),
    ColumnSpec("Analytical Purposes", "purpose_c01", "Purpose C01", 28),
    ColumnSpec("Analytical Purposes", "purpose_c02", "Purpose C02", 28),
    ColumnSpec("Analytical Purposes", "purpose_c03", "Purpose C03", 28),
    ColumnSpec("Analytical Purposes", "purpose_fable", "Purpose Fable", 28),
    ColumnSpec("Analytical Purposes", "purpose_fable_rationale", "Purpose Fable stored rationale", 52),
    ColumnSpec("Analytical Purposes", "purpose_gpt55", "Purpose provisional GPT-5.5", 28),
    ColumnSpec("Analytical Purposes", "purpose_gpt55_rationale", "Purpose provisional GPT-5.5 stored rationale", 52),
    ColumnSpec("C01 diagnostics", "c01_confidence", "C01 confidence", 17),
    ColumnSpec("C01 diagnostics", "c01_register_sufficiency", "C01 register sufficiency", 21),
    ColumnSpec("C01 diagnostics", "c01_taxonomy_fit", "C01 taxonomy fit", 18),
    ColumnSpec("C01 diagnostics", "c01_taxonomy_issue_type", "C01 taxonomy issue type", 28),
    ColumnSpec("C01 diagnostics", "c01_comments", "C01 submitted comments and notes", 52),
    ColumnSpec("C02 diagnostics", "c02_confidence", "C02 confidence", 17),
    ColumnSpec("C02 diagnostics", "c02_register_sufficiency", "C02 register sufficiency", 21),
    ColumnSpec("C02 diagnostics", "c02_taxonomy_fit", "C02 taxonomy fit", 18),
    ColumnSpec("C02 diagnostics", "c02_taxonomy_issue_type", "C02 taxonomy issue type", 28),
    ColumnSpec("C02 diagnostics", "c02_comments", "C02 submitted comments and notes", 52),
    ColumnSpec("C03 diagnostics", "c03_confidence", "C03 confidence", 17),
    ColumnSpec("C03 diagnostics", "c03_register_sufficiency", "C03 register sufficiency", 21),
    ColumnSpec("C03 diagnostics", "c03_taxonomy_fit", "C03 taxonomy fit", 18),
    ColumnSpec("C03 diagnostics", "c03_taxonomy_issue_type", "C03 taxonomy issue type", 28),
    ColumnSpec("C03 diagnostics", "c03_comments", "C03 submitted comments and notes", 52),
    ColumnSpec("Neutral analytical flags", "domain_disagreement_flag", "Domain disagreement flag", 18),
    ColumnSpec("Neutral analytical flags", "purpose_disagreement_flag", "Purpose disagreement flag", 18),
    ColumnSpec("Neutral analytical flags", "any_diagnostic_disagreement_flag", "Any diagnostic disagreement flag", 21),
    ColumnSpec("Neutral analytical flags", "all_human_distinct_flag", "All-human-distinct flag", 18),
    ColumnSpec("Neutral analytical flags", "number_red_human_cells", "Number of red human cells", 16),
    ColumnSpec("Neutral analytical flags", "number_amber_human_cells", "Number of amber human cells", 16),
    ColumnSpec("Neutral analytical flags", "known_instrument_warning", "Known instrument warning", 45),
)


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _resolve(path: Path, root: Path) -> Path:
    return path if path.is_absolute() else root / path


def _relative(path: Path, root: Path) -> str:
    try:
        return path.resolve().relative_to(root.resolve()).as_posix()
    except ValueError:
        return str(path.resolve())


def _git_head(root: Path) -> str:
    result = subprocess.run(
        ["git", "rev-parse", "HEAD"], cwd=root, check=True,
        capture_output=True, text=True,
    )
    return result.stdout.strip()


def _is_git_ignored(path: Path, root: Path) -> bool:
    result = subprocess.run(
        ["git", "check-ignore", "-q", str(path)], cwd=root,
        check=False, capture_output=True, text=True,
    )
    return result.returncode == 0


def validate_restricted_output_dir(output_dir: Path, root: Path) -> None:
    restricted = (root / "preregistration_restricted").resolve()
    resolved = output_dir.resolve()
    if not resolved.is_relative_to(restricted):
        raise CaseReviewError("All data-bearing outputs must be under preregistration_restricted")
    probe = output_dir / "pilot_private_case_review.xlsx"
    if not _is_git_ignored(probe, root):
        raise CaseReviewError(f"Restricted output is not Git-ignored: {probe}")


def _parse_labels(
    value: str, *, delimiter: str, allowed: frozenset[str], context: str
) -> frozenset[str]:
    labels = frozenset(part.strip() for part in str(value).split(delimiter) if part.strip())
    if not labels:
        raise CaseReviewError(f"Missing classification: {context}")
    unknown = labels - allowed
    if unknown:
        raise CaseReviewError(f"Unknown labels in {context}: {sorted(unknown)!r}")
    if UNCLEAR in labels and len(labels) > 1:
        raise CaseReviewError(f"Unclear is combined with substantive labels: {context}")
    return labels


def _canonical_labels(dimension: str, labels: frozenset[str]) -> str:
    order = DOMAIN_ORDER if dimension == "Research Domains" else PURPOSE_ORDER
    rank = {label: index for index, label in enumerate(order)}
    return "\n".join(sorted(labels, key=rank.__getitem__))


def _pattern_text(pattern: str) -> str:
    return {
        "unanimous": "All three identical",
        "all_agree": "All three identical",
        "two_vs_one": "Exactly two identical",
        "all_sets_distinct": "All three distinct",
        "split": "All three distinct",
    }.get(pattern, pattern)


def _human_cell_colors(values: Mapping[str, object], pattern: str) -> dict[str, str]:
    if set(values) != set(CODERS):
        raise CaseReviewError("Human agreement requires C01, C02 and C03")
    counts = Counter(values.values())
    if pattern in {"unanimous", "all_agree"} and list(counts.values()) == [3]:
        return {coder: "green" for coder in CODERS}
    if pattern == "two_vs_one" and sorted(counts.values()) == [1, 2]:
        majority = next(value for value, count in counts.items() if count == 2)
        return {coder: "green" if values[coder] == majority else "red" for coder in CODERS}
    if pattern in {"all_sets_distinct", "split"} and len(counts) == 3:
        return {coder: "amber" for coder in CODERS}
    raise CaseReviewError(f"Agreement pattern {pattern!r} conflicts with values")


def _majority(values: Mapping[str, object]) -> object | None:
    value, count = Counter(values.values()).most_common(1)[0]
    return value if count >= 2 else None


def _model_cell_state(
    model_value: frozenset[str], human_values: Mapping[str, frozenset[str]]
) -> tuple[str, str]:
    majority = _majority(human_values)
    if majority is not None:
        return ("green" if model_value == majority else "red", "")
    matches = [coder for coder in CODERS if model_value == human_values[coder]]
    match_text = ", ".join(matches) if matches else "none"
    return "amber", f"All three human sets differ; exact model match: {match_text}."


def _render_rationale(row: Mapping[str, str], field: str | None) -> str:
    if not field or not row.get(field, "").strip():
        return MISSING_RATIONALE
    raw = row[field]
    try:
        structured = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        rendered = raw
    else:
        if isinstance(structured, (dict, list)):
            rendered = json.dumps(structured, ensure_ascii=False, indent=2, sort_keys=True)
        else:
            rendered = raw
    return (
        f'Overall stored rationale (source field: "{field}"; identical stored text '
        f'is shown in both dimension columns):\n{rendered}'
    )


def _prompt_text(title: str, datasets: str) -> str:
    def sanitise(value: str) -> str:
        text = " ".join(str(value).split()).replace("```", "'''")
        return text.replace("{", "(").replace("}", ")").replace("[", "(").replace("]", ")").strip()

    clean_datasets = " ".join(str(datasets).split())
    if not clean_datasets:
        clean_datasets = "(no datasets listed)"
    elif len(clean_datasets) > 600:
        clean_datasets = clean_datasets[:597] + "..."
    return (
        f"Project title supplied to model:\n{sanitise(title)}\n\n"
        f"Datasets used supplied to model:\n{sanitise(clean_datasets)}"
    )


def load_human_inputs(
    ready_path: Path,
    agreement_path: Path,
    dimension_summary_path: Path,
    raw_path: Path,
) -> tuple[
    list[str],
    dict[tuple[str, str], dict[str, str]],
    dict[tuple[str, str], str],
    dict[tuple[str, str], str],
    dict[tuple[str, str], str],
    Counter[str],
]:
    ready_rows = _read_csv(ready_path)
    if len(ready_rows) != 30:
        raise CaseReviewError(f"Expected 30 human rows; found {len(ready_rows)}")
    ready: dict[tuple[str, str], dict[str, str]] = {}
    for row in ready_rows:
        key = (row["record_id"].strip(), row["coder_id"].strip().upper())
        if key in ready or key[1] not in CODERS:
            raise CaseReviewError(f"Invalid or duplicate human row: {key!r}")
        ready[key] = row
    record_ids = sorted({record_id for record_id, _ in ready})
    expected = {(record_id, coder) for record_id in record_ids for coder in CODERS}
    if len(record_ids) != 10 or set(ready) != expected:
        raise CaseReviewError("Human matrix must contain exactly ten records x three coders")

    patterns: dict[tuple[str, str], str] = {}
    for row in _read_csv(agreement_path):
        dimension = row.get("classification_dimension", "")
        record_id = row.get("record_id", "")
        if record_id not in record_ids or dimension not in CLASSIFICATION_DIMENSIONS:
            continue
        field = "research_domains" if dimension == "Research Domains" else "analytical_purposes"
        allowed = DOMAIN_LABELS if dimension == "Research Domains" else PURPOSE_LABELS
        values = {
            coder: _parse_labels(
                ready[(record_id, coder)][field], delimiter="|", allowed=allowed,
                context=f"{record_id}/{coder}/{dimension}",
            )
            for coder in CODERS
        }
        derived = complete_set_pattern(values.values())
        if row["complete_set_pattern"] != derived:
            raise CaseReviewError(f"Stored agreement pattern mismatch for {record_id}/{dimension}")
        for coder in CODERS:
            stored = _parse_labels(
                row[f"{coder}_set"], delimiter="|", allowed=allowed,
                context=f"agreement/{record_id}/{coder}/{dimension}",
            )
            if stored != values[coder]:
                raise CaseReviewError(f"Stored human labels mismatch for {record_id}/{coder}")
        patterns[(record_id, dimension)] = derived
    if len(patterns) != 20:
        raise CaseReviewError("Classification agreement table is incomplete")

    diagnostic_patterns: dict[tuple[str, str], str] = {}
    for row in _read_csv(dimension_summary_path):
        key = (row.get("record_id", ""), row.get("dimension", ""))
        if key[0] in record_ids and key[1] in DIAGNOSTIC_DIMENSIONS:
            diagnostic_patterns[key] = row["agreement_pattern"]
    if len(diagnostic_patterns) != 30:
        raise CaseReviewError("Diagnostic dimension summary is incomplete")
    diagnostic_fields = {
        "Confidence": "confidence",
        "Register sufficiency": "register_sufficiency",
        "Taxonomy fit": "taxonomy_fit",
    }
    for (record_id, dimension), pattern in diagnostic_patterns.items():
        values = {coder: ready[(record_id, coder)][diagnostic_fields[dimension]] for coder in CODERS}
        _human_cell_colors(values, pattern)

    raw_rows = _read_csv(raw_path)
    if len(raw_rows) != 30:
        raise CaseReviewError(f"Expected 30 raw pilot rows; found {len(raw_rows)}")
    comments: dict[tuple[str, str], str] = {}
    comment_counts: Counter[str] = Counter()
    for row in raw_rows:
        record_id = row.get("source_record_id", "").strip()
        coder = row.get("reviewer_id", "").strip().upper()
        key = (record_id, coder)
        if key not in ready or key in comments:
            raise CaseReviewError(f"Raw reviewer mapping is missing or duplicated: {key!r}")
        note = row.get("sc_note", "")
        note_present = bool(note.strip())
        if note_present:
            note_text = note
            comment_counts[coder] += 1
        elif ready[key]["note_triggered"] == "1":
            note_text = "Blank response (field triggered)"
        else:
            note_text = "Not triggered"
        exposure_note = row.get("sc_exposure_note", "")
        exposure = row.get("sc_exposure", "").strip()
        if exposure_note.strip():
            exposure_text = exposure_note
            comment_counts[coder] += 1
        elif exposure == "1":
            exposure_text = "Blank response (field triggered)"
        elif exposure == "0":
            exposure_text = "Not triggered"
        else:
            exposure_text = "Blank response (trigger state unavailable)"
        if int(note_present) != int(ready[key]["note_present"]):
            raise CaseReviewError(f"Raw/ready note-presence mismatch: {key!r}")
        comments[key] = (
            f"Classification note [sc_note]:\n{note_text}\n\n"
            f"Exposure note [sc_exposure_note]:\n{exposure_text}"
        )
    if set(comments) != expected:
        raise CaseReviewError("Raw comments do not map one-to-one to all coder-record rows")
    return record_ids, ready, patterns, diagnostic_patterns, comments, comment_counts


def load_model_source(
    path: Path,
    *,
    expected_record_ids: set[str],
    source: str,
) -> ModelSource:
    raw_bytes = path.read_bytes()
    encoding = "UTF-8 with BOM" if raw_bytes.startswith(b"\xef\xbb\xbf") else "UTF-8 without BOM"
    try:
        raw_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise CaseReviewError(f"{source} source is not valid UTF-8: {path}") from exc
    rows = _read_csv(path)
    required = {"Record ID", "substantive_domains", "analytical_purpose"}
    if source == "GPT-5.5":
        required |= {"gpt_status", "validation_error"}
    headers = set(rows[0]) if rows else set()
    if not required <= headers:
        raise CaseReviewError(f"{source} source lacks columns: {sorted(required - headers)!r}")
    ids = [row["Record ID"].strip() for row in rows]
    if any(not record_id for record_id in ids):
        raise CaseReviewError(f"{source} contains blank Record IDs")
    duplicates = [record_id for record_id, count in Counter(ids).items() if count > 1]
    if duplicates:
        raise CaseReviewError(f"{source} contains duplicate Record IDs: {duplicates[:5]!r}")
    selected = {row["Record ID"].strip(): row for row in rows if row["Record ID"].strip() in expected_record_ids}
    if set(selected) != expected_record_ids:
        raise CaseReviewError(f"{source} pilot join mismatch: missing {sorted(expected_record_ids - set(selected))!r}")
    for record_id, row in selected.items():
        if source == "GPT-5.5" and (
            row.get("gpt_status", "") != "ok" or row.get("validation_error", "").strip()
        ):
            raise CaseReviewError(f"Invalid stored GPT-5.5 row for {record_id}")
        _parse_labels(
            row["substantive_domains"], delimiter=";", allowed=DOMAIN_LABELS,
            context=f"{source}/{record_id}/Research Domains",
        )
        _parse_labels(
            row["analytical_purpose"], delimiter=";", allowed=PURPOSE_LABELS,
            context=f"{source}/{record_id}/Analytical Purposes",
        )
    rationale_field = "rationale" if "rationale" in headers else None
    return ModelSource(
        path=path,
        sha256=hashlib.sha256(raw_bytes).hexdigest(),
        byte_size=len(raw_bytes),
        row_count=len(rows),
        unique_record_count=len(set(ids)),
        encoding=encoding,
        rows_by_id=selected,
        rationale_field=rationale_field,
    )


def verify_gpt_snapshot(source: ModelSource, pilot_ids: set[str]) -> dict[str, object]:
    if source.row_count not in {1308, 1309}:
        raise CaseReviewError(f"GPT source must have 1,308 or 1,309 rows; found {source.row_count}")
    if source.unique_record_count != source.row_count:
        raise CaseReviewError("GPT source does not have one unique Record ID per row")
    all_rows = _read_csv(source.path)
    all_ids = Counter(row["Record ID"].strip() for row in all_rows)
    variants = {variant: all_ids[variant] for variant in ("2023/211/a", "2023/211/b")}
    if source.row_count == 1309 and variants != {"2023/211/a": 1, "2023/211/b": 1}:
        raise CaseReviewError(f"Pre-collapse duplicate variants are not both present once: {variants}")
    if any(all_ids[record_id] != 1 for record_id in pilot_ids):
        raise CaseReviewError("At least one pilot Record ID is missing or duplicated in GPT source")
    if pilot_ids & set(variants):
        raise CaseReviewError("A pilot record is affected by the 2023/211 variant condition")
    is_provisional = source.path.name == PROVISIONAL_FILENAME
    if is_provisional and source.row_count != 1309:
        raise CaseReviewError("The named provisional snapshot must contain 1,309 rows")
    return {
        "is_provisional": is_provisional,
        "variant_counts": variants,
        "pilot_records_unaffected": True,
    }


def verify_model_provenance(
    fable_metadata_path: Path, gpt_run_script_path: Path, fable: ModelSource
) -> dict[str, str]:
    metadata = json.loads(fable_metadata_path.read_text(encoding="utf-8"))
    expected_fable = {
        "model": "claude-fable-5",
        "run_type": "validation_release_fable5_full_register",
        "prompt_version": EXPECTED_TAXONOMY,
        "taxonomy_version": EXPECTED_TAXONOMY,
        "n_projects": fable.row_count,
    }
    for key, expected in expected_fable.items():
        if metadata.get(key) != expected:
            raise CaseReviewError(f"Fable metadata mismatch for {key}: {metadata.get(key)!r}")
    script = gpt_run_script_path.read_text(encoding="utf-8")
    required_fragments = (
        'MODEL = "gpt-5.5"',
        'TAXONOMY_VERSION = "dict-1.0-rc2"',
        '"run_type": "cross_model_hard_case_disagreement_stratum_not_release"',
        'PROMPT_VERSION = clf.PROMPT_VERSION',
    )
    missing = [fragment for fragment in required_fragments if fragment not in script]
    if missing:
        raise CaseReviewError(f"Archived GPT run evidence lacks expected metadata: {missing!r}")
    classifier = (gpt_run_script_path.parents[1] / "llm_theme_analysis_v3.py").read_text(encoding="utf-8")
    if f'PROMPT_VERSION = "{EXPECTED_TAXONOMY}"' not in classifier:
        raise CaseReviewError("Archived GPT prompt pointer does not resolve to dict-1.0-rc2")
    return {
        "fable_model": str(metadata["model"]),
        "fable_run": str(metadata["run_type"]),
        "fable_prompt": str(metadata["prompt_version"]),
        "fable_taxonomy": str(metadata["taxonomy_version"]),
        "gpt_model": "gpt-5.5",
        "gpt_run": "cross_model_hard_case_disagreement_stratum_not_release",
        "gpt_prompt": EXPECTED_TAXONOMY,
        "gpt_taxonomy": EXPECTED_TAXONOMY,
        "gpt_metadata_basis": "archived run script; the provisional CSV has no embedded version columns",
    }


def compare_gpt_reference(
    reference_path: Path,
    gpt: ModelSource,
    pilot_ids: Sequence[str],
) -> str:
    if not reference_path.exists():
        return f"Not performed: prior restricted comparison file not present at {_relative(reference_path, REPOSITORY_ROOT)}"
    rows = _read_csv(reference_path)
    expected: dict[tuple[str, str], frozenset[str]] = {}
    for row in rows:
        if row.get("source") != "GPT-5.5" or row.get("record_id") not in pilot_ids:
            continue
        dimension = row.get("dimension", "")
        if dimension not in CLASSIFICATION_DIMENSIONS:
            continue
        allowed = DOMAIN_LABELS if dimension == "Research Domains" else PURPOSE_LABELS
        expected[(row["record_id"], dimension)] = _parse_labels(
            row.get("canonical_labels", ""), delimiter="|", allowed=allowed,
            context=f"reference/{row['record_id']}/{dimension}",
        )
    required = {(record_id, dimension) for record_id in pilot_ids for dimension in CLASSIFICATION_DIMENSIONS}
    if set(expected) != required:
        raise CaseReviewError("Existing restricted GPT comparison does not contain one row per pilot dimension")
    for record_id, dimension in sorted(required):
        field = "substantive_domains" if dimension == "Research Domains" else "analytical_purpose"
        allowed = DOMAIN_LABELS if dimension == "Research Domains" else PURPOSE_LABELS
        observed = _parse_labels(
            gpt.rows_by_id[record_id][field], delimiter=";", allowed=allowed,
            context=f"GPT/{record_id}/{dimension}",
        )
        if observed != expected[(record_id, dimension)]:
            raise CaseReviewError(f"Scientific GPT classification difference for {record_id}/{dimension}")
    return "Passed: all ten pilot Domain and Purpose sets exactly match the prior restricted comparison"


def pilot_agreement_summary(
    record_ids: Sequence[str],
    ready: Mapping[tuple[str, str], Mapping[str, str]],
    fable: ModelSource,
    gpt: ModelSource,
) -> list[str]:
    """Return descriptive exact-set counts, never correctness or accuracy claims."""

    lines: list[str] = []
    for source_name, source in (("Fable", fable), ("GPT-5.5", gpt)):
        for dimension, human_field, model_field, allowed in (
            ("Research Domains", "research_domains", "substantive_domains", DOMAIN_LABELS),
            ("Analytical Purposes", "analytical_purposes", "analytical_purpose", PURPOSE_LABELS),
        ):
            coder_matches = Counter({coder: 0 for coder in CODERS})
            majority_available = 0
            majority_matches = 0
            distinct_annotations: list[str] = []
            for record_id in record_ids:
                humans = {
                    coder: _parse_labels(
                        ready[(record_id, coder)][human_field], delimiter="|", allowed=allowed,
                        context=f"summary/{record_id}/{coder}/{dimension}",
                    )
                    for coder in CODERS
                }
                model = _parse_labels(
                    source.rows_by_id[record_id][model_field], delimiter=";", allowed=allowed,
                    context=f"summary/{source_name}/{record_id}/{dimension}",
                )
                for coder in CODERS:
                    coder_matches[coder] += model == humans[coder]
                majority = _majority(humans)
                if majority is not None:
                    majority_available += 1
                    majority_matches += model == majority
                else:
                    matches = [coder for coder in CODERS if model == humans[coder]]
                    distinct_annotations.append(
                        f"{record_id}={'/'.join(matches) if matches else 'none'}"
                    )
            lines.append(
                f"{source_name} / {dimension}: exact C01={coder_matches['C01']}/10, "
                f"C02={coder_matches['C02']}/10, C03={coder_matches['C03']}/10; "
                f"human-majority={majority_matches}/{majority_available}; "
                f"all-human-distinct annotations={', '.join(distinct_annotations) or 'none'}"
            )
    for dimension, model_field, allowed in (
        ("Research Domains", "substantive_domains", DOMAIN_LABELS),
        ("Analytical Purposes", "analytical_purpose", PURPOSE_LABELS),
    ):
        matches = 0
        for record_id in record_ids:
            left = _parse_labels(
                fable.rows_by_id[record_id][model_field], delimiter=";", allowed=allowed,
                context=f"summary/Fable/{record_id}/{dimension}",
            )
            right = _parse_labels(
                gpt.rows_by_id[record_id][model_field], delimiter=";", allowed=allowed,
                context=f"summary/GPT/{record_id}/{dimension}",
            )
            matches += left == right
        lines.append(f"Fable / GPT-5.5 exact-set agreement / {dimension}: {matches}/10")
    return lines


def _build_case_rows(
    record_ids: Sequence[str],
    ready: Mapping[tuple[str, str], Mapping[str, str]],
    patterns: Mapping[tuple[str, str], str],
    diagnostic_patterns: Mapping[tuple[str, str], str],
    comments: Mapping[tuple[str, str], str],
    fable: ModelSource,
    gpt: ModelSource,
) -> tuple[list[dict[str, object]], dict[tuple[str, str], str], dict[tuple[str, str], str]]:
    output: list[dict[str, object]] = []
    cell_colors: dict[tuple[str, str], str] = {}
    cell_annotations: dict[tuple[str, str], str] = {}
    for record_id in record_ids:
        fable_row = fable.rows_by_id[record_id]
        gpt_row = gpt.rows_by_id[record_id]
        row: dict[str, object] = {
            "record_id": record_id,
            "project_title": fable_row["Title"],
            "public_register_text": _prompt_text(fable_row["Title"], fable_row.get("Datasets Used", "")),
        }
        red_count = 0
        amber_count = 0
        all_distinct_dimensions: list[str] = []
        model_mismatches = 0
        for dimension, prefix, human_field, model_field, allowed in (
            ("Research Domains", "domain", "research_domains", "substantive_domains", DOMAIN_LABELS),
            ("Analytical Purposes", "purpose", "analytical_purposes", "analytical_purpose", PURPOSE_LABELS),
        ):
            human_sets = {
                coder: _parse_labels(
                    ready[(record_id, coder)][human_field], delimiter="|", allowed=allowed,
                    context=f"human/{record_id}/{coder}/{dimension}",
                )
                for coder in CODERS
            }
            pattern = patterns[(record_id, dimension)]
            row[f"{prefix}_agreement_pattern"] = _pattern_text(pattern)
            colors = _human_cell_colors(human_sets, pattern)
            for coder in CODERS:
                key = f"{prefix}_{coder.lower()}"
                row[key] = _canonical_labels(dimension, human_sets[coder])
                cell_colors[(record_id, key)] = colors[coder]
                red_count += colors[coder] == "red"
                amber_count += colors[coder] == "amber"
            if pattern == "all_sets_distinct":
                all_distinct_dimensions.append(dimension)
            for model_name, model_source in (("fable", fable), ("gpt55", gpt)):
                labels = _parse_labels(
                    model_source.rows_by_id[record_id][model_field], delimiter=";", allowed=allowed,
                    context=f"{model_name}/{record_id}/{dimension}",
                )
                key = f"{prefix}_{model_name}"
                row[key] = _canonical_labels(dimension, labels)
                color, annotation = _model_cell_state(labels, human_sets)
                cell_colors[(record_id, key)] = color
                if annotation:
                    cell_annotations[(record_id, key)] = annotation
                if _majority(human_sets) is not None and labels != _majority(human_sets):
                    model_mismatches += 1
                row[f"{prefix}_{model_name}_rationale"] = _render_rationale(
                    model_source.rows_by_id[record_id], model_source.rationale_field
                )
        diagnostic_disagreements = 0
        for dimension, field, suffix in (
            ("Confidence", "confidence", "confidence"),
            ("Register sufficiency", "register_sufficiency", "register_sufficiency"),
            ("Taxonomy fit", "taxonomy_fit", "taxonomy_fit"),
        ):
            values = {coder: ready[(record_id, coder)][field] for coder in CODERS}
            pattern = diagnostic_patterns[(record_id, dimension)]
            colors = _human_cell_colors(values, pattern)
            diagnostic_disagreements += pattern != "all_agree"
            for coder in CODERS:
                key = f"{coder.lower()}_{suffix}"
                row[key] = values[coder]
                cell_colors[(record_id, key)] = colors[coder]
                red_count += colors[coder] == "red"
                amber_count += colors[coder] == "amber"
        for coder in CODERS:
            issues = ready[(record_id, coder)]["taxonomy_issues"]
            row[f"{coder.lower()}_taxonomy_issue_type"] = issues.replace("|", "\n") if issues else "Not applicable / no issue selected"
            row[f"{coder.lower()}_comments"] = comments[(record_id, coder)]
        warning = KNOWN_WARNING if record_id == "2021/103" else ""
        row.update({
            "domain_disagreement_flag": int(patterns[(record_id, "Research Domains")] != "unanimous"),
            "purpose_disagreement_flag": int(patterns[(record_id, "Analytical Purposes")] != "unanimous"),
            "any_diagnostic_disagreement_flag": int(diagnostic_disagreements > 0),
            "all_human_distinct_flag": int(bool(all_distinct_dimensions)),
            "number_red_human_cells": red_count,
            "number_amber_human_cells": amber_count,
            "known_instrument_warning": warning,
            "_all_human_distinct_dimensions": "; ".join(all_distinct_dimensions) or "None",
            "_diagnostic_disagreement_count": diagnostic_disagreements,
            "_human_model_mismatch_count": model_mismatches,
            "_human_disagreement_count": (
                int(patterns[(record_id, "Research Domains")] != "unanimous")
                + int(patterns[(record_id, "Analytical Purposes")] != "unanimous")
            ),
        })
        output.append(row)
    if len(output) != 10 or len({str(row["record_id"]) for row in output}) != 10:
        raise CaseReviewError("Case review must contain exactly ten unique rows")
    return output, cell_colors, cell_annotations


def _write_csv(path: Path, rows: Sequence[Mapping[str, object]]) -> None:
    keys = [column.key for column in COLUMNS]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=keys, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _case_sheet(
    workbook: Workbook,
    rows: Sequence[Mapping[str, object]],
    cell_colors: Mapping[tuple[str, str], str],
    cell_annotations: Mapping[tuple[str, str], str],
) -> None:
    sheet = workbook.active
    sheet.title = "Case review"
    groups: list[tuple[str, int, int]] = []
    start = 1
    current = COLUMNS[0].group
    for index, column in enumerate(COLUMNS, start=1):
        if column.group != current:
            groups.append((current, start, index - 1))
            current, start = column.group, index
    groups.append((current, start, len(COLUMNS)))
    for group, first, last in groups:
        if first != last:
            sheet.merge_cells(start_row=1, start_column=first, end_row=1, end_column=last)
        cell = sheet.cell(1, first, group)
        cell.fill = PatternFill("solid", fgColor=COLORS["group"])
        cell.font = Font(color=COLORS["white"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_index in range(first, last + 1):
            sheet.cell(1, column_index).fill = PatternFill("solid", fgColor=COLORS["group"])
    for column_index, spec in enumerate(COLUMNS, start=1):
        cell = sheet.cell(2, column_index, spec.label)
        cell.fill = PatternFill("solid", fgColor=COLORS["header"])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.column_dimensions[get_column_letter(column_index)].width = spec.width
    group_starts = {first for _, first, _ in groups if first > 1}
    thin = Side(style="thin", color="FF7F7F7F")
    for row_index, row in enumerate(rows, start=3):
        record_id = str(row["record_id"])
        for column_index, spec in enumerate(COLUMNS, start=1):
            cell = sheet.cell(row_index, column_index, row.get(spec.key, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(name="Calibri", size=10)
            if column_index in group_starts:
                cell.border = Border(left=thin)
            color = cell_colors.get((record_id, spec.key))
            if color:
                cell.fill = PatternFill("solid", fgColor=COLORS[color])
            elif spec.key.endswith("comments") or spec.key.endswith("rationale"):
                cell.fill = PatternFill("solid", fgColor=COLORS["neutral"])
            annotation = cell_annotations.get((record_id, spec.key))
            if annotation:
                cell.comment = Comment(annotation + " Agreement colour does not indicate correctness.", "Offline builder")
        if record_id == "2021/103":
            warning_column = next(i for i, spec in enumerate(COLUMNS, 1) if spec.key == "c02_taxonomy_issue_type")
            warning_cell = sheet.cell(row_index, warning_column)
            warning_cell.fill = PatternFill("solid", fgColor=COLORS["warning"])
            warning_cell.comment = Comment(KNOWN_WARNING, "Offline builder")
            flag_column = next(i for i, spec in enumerate(COLUMNS, 1) if spec.key == "known_instrument_warning")
            sheet.cell(row_index, flag_column).fill = PatternFill("solid", fgColor=COLORS["warning"])
        sheet.row_dimensions[row_index].height = 132
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 48
    sheet.freeze_panes = "D3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{len(rows) + 2}"
    sheet.sheet_view.showGridLines = False


def _index_sheet(workbook: Workbook, rows: Sequence[Mapping[str, object]]) -> None:
    sheet = workbook.create_sheet("Disagreement index")
    headers = (
        "Record ID", "Project title", "Domain pattern", "Purpose pattern",
        "Diagnostic disagreement count", "All-human-distinct dimensions",
        "Human-model mismatch count", "Case review worksheet row",
    )
    for column_index, header in enumerate(headers, 1):
        cell = sheet.cell(1, column_index, header)
        cell.fill = PatternFill("solid", fgColor=COLORS["group"])
        cell.font = Font(color=COLORS["white"], bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    case_rows = {str(row["record_id"]): index for index, row in enumerate(rows, start=3)}
    sorted_rows = sorted(
        rows,
        key=lambda row: (
            -int(row["all_human_distinct_flag"]),
            -int(row["_human_disagreement_count"]),
            str(row["record_id"]),
        ),
    )
    for row_index, row in enumerate(sorted_rows, 2):
        record_id = str(row["record_id"])
        values = (
            record_id,
            row["project_title"],
            row["domain_agreement_pattern"],
            row["purpose_agreement_pattern"],
            row["_diagnostic_disagreement_count"],
            row["_all_human_distinct_dimensions"],
            row["_human_model_mismatch_count"],
            case_rows[record_id],
        )
        for column_index, value in enumerate(values, 1):
            cell = sheet.cell(row_index, column_index, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        link = sheet.cell(row_index, 1)
        link.hyperlink = f"#'Case review'!A{case_rows[record_id]}"
        link.style = "Hyperlink"
    widths = (16, 48, 22, 22, 19, 28, 18, 18)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:H{len(rows) + 1}"
    sheet.sheet_view.showGridLines = False


def _provenance_sheet(
    workbook: Workbook,
    *,
    sources: Sequence[tuple[str, Path, str, int]],
    gpt: ModelSource,
    snapshot: Mapping[str, object],
    provenance: Mapping[str, str],
    git_head: str,
    created_at: str,
    reference_status: str,
    comment_counts: Mapping[str, int],
    scientific_summary: Sequence[str],
) -> None:
    sheet = workbook.create_sheet("Provenance and legend")
    sheet.merge_cells("A1:B1")
    sheet["A1"] = PROVISIONAL_WARNING_TITLE if snapshot["is_provisional"] else "GPT source status: supplied snapshot"
    sheet["A1"].fill = PatternFill("solid", fgColor=COLORS["warning"])
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.merge_cells("A2:B5")
    sheet["A2"] = PROVISIONAL_WARNING if snapshot["is_provisional"] else "The supplied GPT source is not identified as the provisional 1,309-row snapshot."
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet["A2"].fill = PatternFill("solid", fgColor=COLORS["warning"])
    sheet.row_dimensions[2].height = 95
    row_index = 7

    def section(title: str) -> None:
        nonlocal row_index
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
        cell = sheet.cell(row_index, 1, title)
        cell.fill = PatternFill("solid", fgColor=COLORS["group"])
        cell.font = Font(color=COLORS["white"], bold=True)
        row_index += 1

    def item(key: str, value: object) -> None:
        nonlocal row_index
        sheet.cell(row_index, 1, key).font = Font(bold=True)
        sheet.cell(row_index, 2, value)
        for column in (1, 2):
            sheet.cell(row_index, column).alignment = Alignment(wrap_text=True, vertical="top")
        row_index += 1

    section("Generation")
    item("Created at (UTC)", created_at)
    item("Repository HEAD", git_head)
    item("Interpretation", "Agreement colouring is navigation only. It does not identify correctness, errors, preferred classifications, diagnoses or actions.")
    section("Input files")
    for label, path, digest, count in sources:
        item(label, f"{_relative(path, REPOSITORY_ROOT)}\nSHA-256: {digest}\nRows: {count}")
    item("GPT encoding / bytes", f"{gpt.encoding}; {gpt.byte_size} bytes")
    item("GPT duplicate condition", f"2023/211/a={snapshot['variant_counts']['2023/211/a']}; 2023/211/b={snapshot['variant_counts']['2023/211/b']}; ten pilot records unaffected=yes")
    item("Prior restricted GPT comparison", reference_status)
    section("Direct pilot exact-set checks")
    item(
        "Scope",
        "Descriptive agreement checks only; these are not accuracy, correctness, release or formal validation metrics.",
    )
    for index, line in enumerate(scientific_summary, 1):
        item(f"Check {index}", line)
    section("Taxonomy and model provenance")
    item("Taxonomy / prompt", EXPECTED_TAXONOMY)
    item("Fable", f"Model: {provenance['fable_model']}; run: {provenance['fable_run']}")
    item("GPT-5.5", f"Model: {provenance['gpt_model']}; run: {provenance['gpt_run']}\nBasis: {provenance['gpt_metadata_basis']}")
    section("Comment and rationale provenance")
    item("Coder mapping", "Raw hidden reviewer_id mapped directly to C01/C02/C03; no row-order or timestamp inference")
    item("Coder fields", "sc_note (conditional classification note); sc_exposure_note (conditional prohibited-exposure note). po_note excluded as project-owner material. Blank and not-triggered states remain distinct.")
    item("Submitted comment counts", "; ".join(f"{coder}={comment_counts.get(coder, 0)}" for coder in CODERS))
    item("Fable rationale", f"Source field: {provenance.get('fable_rationale_field', 'not available')}; one overall stored rationale, shown with an explicit duplication label in both dimension columns")
    item("GPT rationale", f"Source field: {provenance.get('gpt_rationale_field', 'not available')}; one overall stored rationale, shown with an explicit duplication label in both dimension columns")
    section("Colour legend")
    item("Green", "Human: unanimous or agreeing pair. Model: exact match to an available complete-set human majority.")
    item("Red", "Human: lone dissenter. Model: differs from an available complete-set human majority.")
    item("Amber", "All three human values differ; model cells are annotated with exact coder-set matches or none.")
    item("Warning fill", KNOWN_WARNING)
    section("Column definitions")
    for spec in COLUMNS:
        item(spec.key, f"{spec.group}: {spec.label}")
    sheet.column_dimensions["A"].width = 36
    sheet.column_dimensions["B"].width = 110
    sheet.sheet_view.showGridLines = False


def _write_legend(
    path: Path,
    *,
    sources: Sequence[tuple[str, Path, str, int]],
    gpt: ModelSource,
    snapshot: Mapping[str, object],
    provenance: Mapping[str, str],
    git_head: str,
    created_at: str,
    reference_status: str,
    comment_counts: Mapping[str, int],
    scientific_summary: Sequence[str],
) -> None:
    source_lines = "\n".join(
        f"- {label}: `{_relative(source_path, REPOSITORY_ROOT)}`; SHA-256 `{digest}`; rows `{count}`."
        for label, source_path, digest, count in sources
    )
    warning = f"## {PROVISIONAL_WARNING_TITLE}\n\n{PROVISIONAL_WARNING}" if snapshot["is_provisional"] else "## GPT source status\n\nSupplied replaceable GPT snapshot."
    text = f"""# Restricted pilot private case-review legend

{warning}

- Actual GPT SHA-256: `{gpt.sha256}`.
- GPT byte size: `{gpt.byte_size}`; encoding: `{gpt.encoding}`; data rows: `{gpt.row_count}`.
- Pre-collapse condition: `2023/211/a` and `2023/211/b` each occur once.
- The ten pilot Record IDs each occur once and none is affected by that condition.
- Repository HEAD: `{git_head}`.
- Generated at: `{created_at}`.

## Sources

{source_lines}

Prior restricted GPT comparison check: {reference_status}

## Direct pilot exact-set checks

These are descriptive agreement checks, not correctness, accuracy, release or formal validation metrics.

{chr(10).join(f'- {line}' for line in scientific_summary)}

## Provenance

- Taxonomy and prompt: `{EXPECTED_TAXONOMY}`.
- Fable: `{provenance['fable_model']}` / `{provenance['fable_run']}`.
- Provisional GPT: `{provenance['gpt_model']}` / `{provenance['gpt_run']}`.
- GPT version basis: {provenance['gpt_metadata_basis']}.
- Public-register context reproduces the production prompt inputs: cleaned `Title` plus the deterministically sanitised and at-most-600-character `Datasets Used` field. It is not a new summary.

## Comments and rationales

- Coder mapping uses the hidden raw `reviewer_id` directly for C01/C02/C03.
- Comment fields: `sc_note` and `sc_exposure_note`. Source labels and original submitted wording are retained; triggered blanks and non-triggered fields are distinguished.
- Submitted non-empty comment counts: {', '.join(f'{coder}={comment_counts.get(coder, 0)}' for coder in CODERS)}.
- Fable rationale field: `{provenance.get('fable_rationale_field', 'not available')}`.
- GPT rationale field: `{provenance.get('gpt_rationale_field', 'not available')}`.
- Each model source stores one overall rationale rather than dimension-specific rationales. The same verbatim rationale is shown in both dimension columns with an explicit duplication label. Missing text is displayed as `{MISSING_RATIONALE}`.

## Agreement colours

- Green: unanimous human values or the agreeing human pair; for a model, exact match to an available complete-set human majority.
- Red: lone human dissenter; for a model, difference from an available complete-set human majority.
- Amber: all three human values differ. Model cells state whether they exactly match C01, C02, C03 or none.
- Warning fill: {KNOWN_WARNING}

Colours represent agreement only, not correctness. The workbook makes no final diagnosis, preferred classification, formal validation estimate, release decision or replacement-panel assessment.

The CSV is a restricted machine-readable companion and does not preserve Excel formatting or cell comments.
"""
    path.write_text(text, encoding="utf-8")


def build_private_case_review(
    *,
    gpt_source: Path,
    output_dir: Path,
    ready_path: Path = DEFAULT_READY,
    agreement_path: Path = DEFAULT_AGREEMENT,
    dimension_summary_path: Path = DEFAULT_DIMENSION_SUMMARY,
    raw_path: Path = DEFAULT_RAW,
    fable_path: Path = DEFAULT_FABLE,
    fable_metadata_path: Path = DEFAULT_FABLE_METADATA,
    gpt_run_script_path: Path = DEFAULT_GPT_RUN_SCRIPT,
    gpt_reference_path: Path = DEFAULT_GPT_REFERENCE,
    root: Path = REPOSITORY_ROOT,
) -> dict[str, object]:
    resolved = {
        "gpt": _resolve(gpt_source, root),
        "output": _resolve(output_dir, root),
        "ready": _resolve(ready_path, root),
        "agreement": _resolve(agreement_path, root),
        "dimension": _resolve(dimension_summary_path, root),
        "raw": _resolve(raw_path, root),
        "fable": _resolve(fable_path, root),
        "fable_metadata": _resolve(fable_metadata_path, root),
        "gpt_run_script": _resolve(gpt_run_script_path, root),
        "gpt_reference": _resolve(gpt_reference_path, root),
        "builder": Path(__file__).resolve(),
    }
    validate_restricted_output_dir(resolved["output"], root)
    source_paths = [
        resolved[key]
        for key in (
            "gpt", "ready", "agreement", "dimension", "raw", "fable",
            "fable_metadata", "gpt_run_script", "builder",
        )
    ]
    missing = [str(path) for path in source_paths if not path.is_file()]
    if missing:
        raise CaseReviewError(f"Missing source files: {missing!r}")
    before_hashes = {path: _sha256(path) for path in source_paths}
    record_ids, ready, patterns, diagnostic_patterns, comments, comment_counts = load_human_inputs(
        resolved["ready"], resolved["agreement"], resolved["dimension"], resolved["raw"]
    )
    fable = load_model_source(
        resolved["fable"], expected_record_ids=set(record_ids), source="Fable"
    )
    gpt = load_model_source(
        resolved["gpt"], expected_record_ids=set(record_ids), source="GPT-5.5"
    )
    snapshot = verify_gpt_snapshot(gpt, set(record_ids))
    provenance = verify_model_provenance(
        resolved["fable_metadata"], resolved["gpt_run_script"], fable
    )
    provenance["fable_rationale_field"] = fable.rationale_field or "not available"
    provenance["gpt_rationale_field"] = gpt.rationale_field or "not available"
    reference_status = compare_gpt_reference(resolved["gpt_reference"], gpt, record_ids)
    scientific_summary = pilot_agreement_summary(record_ids, ready, fable, gpt)
    rows, cell_colors, cell_annotations = _build_case_rows(
        record_ids, ready, patterns, diagnostic_patterns, comments, fable, gpt
    )
    created_at = datetime.now(timezone.utc).isoformat()
    git_head = _git_head(root)
    sources = (
        ("Human analysis-ready", resolved["ready"], before_hashes[resolved["ready"]], len(_read_csv(resolved["ready"]))),
        ("Human record agreement", resolved["agreement"], before_hashes[resolved["agreement"]], len(_read_csv(resolved["agreement"]))),
        ("Human dimension summary", resolved["dimension"], before_hashes[resolved["dimension"]], len(_read_csv(resolved["dimension"]))),
        ("Raw pilot export", resolved["raw"], before_hashes[resolved["raw"]], len(_read_csv(resolved["raw"]))),
        ("Fable classifications", resolved["fable"], fable.sha256, fable.row_count),
        ("GPT classifications", resolved["gpt"], gpt.sha256, gpt.row_count),
        ("Fable run metadata", resolved["fable_metadata"], before_hashes[resolved["fable_metadata"]], 1),
        ("Archived GPT run script", resolved["gpt_run_script"], before_hashes[resolved["gpt_run_script"]], 1),
        ("Workbook builder", resolved["builder"], before_hashes[resolved["builder"]], 1),
    )
    resolved["output"].mkdir(parents=True, exist_ok=True)
    csv_path = resolved["output"] / "pilot_private_case_review_wide.csv"
    xlsx_path = resolved["output"] / "pilot_private_case_review.xlsx"
    legend_path = resolved["output"] / "pilot_private_case_review_legend.md"
    _write_csv(csv_path, rows)
    workbook = Workbook()
    _case_sheet(workbook, rows, cell_colors, cell_annotations)
    _index_sheet(workbook, rows)
    _provenance_sheet(
        workbook, sources=sources, gpt=gpt, snapshot=snapshot,
        provenance=provenance, git_head=git_head, created_at=created_at,
        reference_status=reference_status, comment_counts=comment_counts,
        scientific_summary=scientific_summary,
    )
    workbook.save(xlsx_path)
    _write_legend(
        legend_path, sources=sources, gpt=gpt, snapshot=snapshot,
        provenance=provenance, git_head=git_head, created_at=created_at,
        reference_status=reference_status, comment_counts=comment_counts,
        scientific_summary=scientific_summary,
    )
    after_hashes = {path: _sha256(path) for path in source_paths}
    changed = [str(path) for path in source_paths if before_hashes[path] != after_hashes[path]]
    if changed:
        raise CaseReviewError(f"Source files changed during generation: {changed!r}")
    for path in (csv_path, xlsx_path, legend_path):
        if not _is_git_ignored(path, root):
            raise CaseReviewError(f"Generated restricted output is not Git-ignored: {path}")
    return {
        "record_ids": record_ids,
        "rows": rows,
        "gpt": gpt,
        "fable": fable,
        "snapshot": snapshot,
        "reference_status": reference_status,
        "scientific_summary": scientific_summary,
        "comment_counts": dict(comment_counts),
        "outputs": [xlsx_path, csv_path, legend_path],
        "source_hashes": {str(path): digest for path, digest in before_hashes.items()},
    }


def inspect_workbook(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, data_only=False)
    case = workbook["Case review"]
    return {
        "sheet_names": workbook.sheetnames,
        "dimensions": {
            sheet.title: (sheet.max_row, sheet.max_column)
            for sheet in workbook.worksheets
        },
        "case_data_rows": case.max_row - 2,
        "case_freeze_panes": str(case.freeze_panes),
        "case_filter": case.auto_filter.ref,
        "wrapped_data_cells": sum(bool(cell.alignment.wrap_text) for row in case.iter_rows(min_row=3) for cell in row),
        "comment_cells": sum(cell.comment is not None for row in case.iter_rows() for cell in row),
        "warning_present": any(PROVISIONAL_WARNING_TITLE in str(cell.value) for row in workbook["Provenance and legend"].iter_rows() for cell in row if cell.value),
    }


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Build an offline restricted pilot case-review XLSX/CSV/legend."
    )
    parser.add_argument(
        "--gpt-source", required=True, type=Path,
        help="GPT classification CSV to use. Replace this path to regenerate against the canonical snapshot.",
    )
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--ready", type=Path, default=DEFAULT_READY)
    parser.add_argument("--agreement", type=Path, default=DEFAULT_AGREEMENT)
    parser.add_argument("--dimension-summary", type=Path, default=DEFAULT_DIMENSION_SUMMARY)
    parser.add_argument("--raw-pilot", type=Path, default=DEFAULT_RAW)
    parser.add_argument("--fable-source", type=Path, default=DEFAULT_FABLE)
    parser.add_argument("--fable-metadata", type=Path, default=DEFAULT_FABLE_METADATA)
    parser.add_argument("--gpt-run-script", type=Path, default=DEFAULT_GPT_RUN_SCRIPT)
    parser.add_argument("--gpt-reference", type=Path, default=DEFAULT_GPT_REFERENCE)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    result = build_private_case_review(
        gpt_source=args.gpt_source,
        output_dir=args.output_dir,
        ready_path=args.ready,
        agreement_path=args.agreement,
        dimension_summary_path=args.dimension_summary,
        raw_path=args.raw_pilot,
        fable_path=args.fable_source,
        fable_metadata_path=args.fable_metadata,
        gpt_run_script_path=args.gpt_run_script,
        gpt_reference_path=args.gpt_reference,
    )
    gpt: ModelSource = result["gpt"]  # type: ignore[assignment]
    print(f"Created restricted case review for {len(result['record_ids'])} pilot records")
    print(f"GPT source: {_relative(gpt.path, REPOSITORY_ROOT)}")
    print(f"GPT SHA-256: {gpt.sha256}; rows: {gpt.row_count}; bytes: {gpt.byte_size}")
    for path in result["outputs"]:
        print(f"Output: {_relative(path, REPOSITORY_ROOT)}")
    print(f"Prior restricted GPT comparison: {result['reference_status']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
