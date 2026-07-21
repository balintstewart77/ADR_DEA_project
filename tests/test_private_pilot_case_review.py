import csv
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

import analysis.validation.build_private_pilot_case_review as review


RECORD_IDS = [
    "2021/103",
    *[f"SYN-{index:03}" for index in range(1, 10)],
]


def _write_csv(path: Path, columns, rows, *, bom=False):
    path.parent.mkdir(parents=True, exist_ok=True)
    encoding = "utf-8-sig" if bom else "utf-8"
    with path.open("w", encoding=encoding, newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def _human_values(index):
    domains = {coder: "Education & Skills" for coder in review.CODERS}
    purposes = {coder: "Descriptive Monitoring" for coder in review.CODERS}
    confidence = {coder: "High" for coder in review.CODERS}
    sufficiency = {coder: "Sufficient" for coder in review.CODERS}
    fit = {coder: "Fit" for coder in review.CODERS}
    if index == 1:
        domains["C03"] = "Health & Social Care"
        purposes["C03"] = "Outcome Tracking"
        confidence["C03"] = "Low"
    if index == 2:
        domains.update({"C02": "Health & Social Care", "C03": "Crime & Justice"})
        purposes.update({
            "C02": "Outcome Tracking",
            "C03": "Risk Prediction / Early Identification",
        })
        confidence.update({"C02": "Medium", "C03": "Low"})
        fit.update({"C02": "Partial Fit", "C03": "No Fit"})
    return domains, purposes, confidence, sufficiency, fit


def _pattern(values, set_dimension=False):
    unique = len(set(values.values()))
    if unique == 1:
        return "unanimous" if set_dimension else "all_agree"
    if unique == 2:
        return "two_vs_one"
    return "all_sets_distinct" if set_dimension else "split"


def _make_inputs(root: Path):
    ready_rows = []
    agreement_rows = []
    dimension_rows = []
    raw_rows = []
    for index, record_id in enumerate(RECORD_IDS):
        domains, purposes, confidence, sufficiency, fit = _human_values(index)
        for coder in review.CODERS:
            submitted = index == 0
            triggered_blank = index == 1 and coder == "C01"
            taxonomy_issues = "None" if record_id == "2021/103" and coder == "C02" else ""
            ready_rows.append({
                "record_id": record_id,
                "coder_id": coder,
                "instrument_version": "redcap-candidate-0.3",
                "research_domains": domains[coder],
                "analytical_purposes": purposes[coder],
                "covid_tag": "0",
                "equity_tag": "0",
                "register_sufficiency": sufficiency[coder],
                "taxonomy_fit": fit[coder],
                "taxonomy_issues": taxonomy_issues,
                "confidence": confidence[coder],
                "response_complete": "1",
                "note_triggered": "1" if submitted or triggered_blank else "0",
                "note_present": "1" if submitted else "0",
                "qc_status": "valid_response",
            })
            raw_rows.append({
                "source_record_id": record_id,
                "reviewer_id": coder,
                "sc_note": f"synthetic note for {coder}" if submitted else "",
                "sc_exposure": "0",
                "sc_exposure_note": "",
            })
        for dimension, values in (
            ("Research Domains", domains),
            ("Analytical Purposes", purposes),
        ):
            agreement_rows.append({
                "record_id": record_id,
                "classification_dimension": dimension,
                "complete_set_pattern": _pattern(values, True),
                **{f"{coder}_set": values[coder] for coder in review.CODERS},
            })
        for dimension, values in (
            ("Confidence", confidence),
            ("Register sufficiency", sufficiency),
            ("Taxonomy fit", fit),
        ):
            dimension_rows.append({
                "record_id": record_id,
                "dimension": dimension,
                "agreement_pattern": _pattern(values),
            })

    ready_path = root / review.DEFAULT_READY
    agreement_path = root / review.DEFAULT_AGREEMENT
    dimension_path = root / review.DEFAULT_DIMENSION_SUMMARY
    raw_path = root / review.DEFAULT_RAW
    _write_csv(ready_path, list(ready_rows[0]), ready_rows)
    _write_csv(agreement_path, list(agreement_rows[0]), agreement_rows)
    _write_csv(dimension_path, list(dimension_rows[0]), dimension_rows)
    _write_csv(raw_path, list(raw_rows[0]), raw_rows)

    fable_rows = []
    gpt_pilot_rows = []
    for index, record_id in enumerate(RECORD_IDS):
        domain = "Education & Skills"
        purpose = "Descriptive Monitoring"
        if index == 2:
            purpose = "Methodological / Infrastructure Research"
        common = {
            "Record ID": record_id,
            "Title": f"Synthetic project {index}",
            "Datasets Used": f"Synthetic public dataset {index}",
            "substantive_domains": domain,
            "analytical_purpose": purpose,
            "rationale": f"Stored overall rationale {index}",
        }
        fable_rows.append(dict(common))
        gpt_pilot_rows.append({
            **common,
            "gpt_status": "ok",
            "validation_error": "",
        })
    fable_path = root / review.DEFAULT_FABLE
    _write_csv(fable_path, list(fable_rows[0]), fable_rows)
    metadata_path = root / review.DEFAULT_FABLE_METADATA
    metadata_path.parent.mkdir(parents=True, exist_ok=True)
    metadata_path.write_text(json.dumps({
        "model": "claude-fable-5",
        "run_type": "validation_release_fable5_full_register",
        "prompt_version": review.EXPECTED_TAXONOMY,
        "taxonomy_version": review.EXPECTED_TAXONOMY,
        "n_projects": 10,
    }), encoding="utf-8")

    gpt_rows = list(gpt_pilot_rows)
    gpt_rows.append(
        {
            "Record ID": "2023/211",
            "Title": "Synthetic canonical retained record",
            "Datasets Used": "Synthetic",
            "substantive_domains": "Education & Skills",
            "analytical_purpose": "Descriptive Monitoring",
            "rationale": "Stored",
            "gpt_status": "ok",
            "validation_error": "",
        }
    )
    for index in range(1308 - len(gpt_rows)):
        gpt_rows.append({
            "Record ID": f"EXTRA-{index:04}",
            "Title": "Synthetic extra",
            "Datasets Used": "Synthetic",
            "substantive_domains": "Education & Skills",
            "analytical_purpose": "Descriptive Monitoring",
            "rationale": "Stored",
            "gpt_status": "ok",
            "validation_error": "",
        })
    gpt_path = root / review.DEFAULT_GPT_SOURCE
    _write_csv(gpt_path, list(gpt_rows[0]), gpt_rows, bom=True)

    run_script = root / review.DEFAULT_GPT_RUN_SCRIPT
    run_script.parent.mkdir(parents=True, exist_ok=True)
    run_script.write_text(
        'MODEL = "gpt-5.5"\n'
        'PROMPT_VERSION = clf.PROMPT_VERSION\n'
        'TAXONOMY_VERSION = "dict-1.0-rc2"\n'
        'metadata = {"run_type": "cross_model_hard_case_disagreement_stratum_not_release"}\n',
        encoding="utf-8",
    )
    classifier = root / "analysis" / "llm_theme_analysis_v3.py"
    classifier.write_text('PROMPT_VERSION = "dict-1.0-rc2"\n', encoding="utf-8")

    reference_path = root / review.DEFAULT_GPT_REFERENCE
    reference_rows = []
    for row in gpt_pilot_rows:
        for dimension, field in (
            ("Research Domains", "substantive_domains"),
            ("Analytical Purposes", "analytical_purpose"),
        ):
            reference_rows.append({
                "record_id": row["Record ID"],
                "dimension": dimension,
                "source": "GPT-5.5",
                "canonical_labels": row[field],
            })
    _write_csv(reference_path, list(reference_rows[0]), reference_rows)
    return {
        "ready": ready_path,
        "agreement": agreement_path,
        "dimension": dimension_path,
        "raw": raw_path,
        "fable": fable_path,
        "metadata": metadata_path,
        "gpt": gpt_path,
        "run_script": run_script,
        "reference": reference_path,
    }


@pytest.fixture
def built(tmp_path, monkeypatch):
    paths = _make_inputs(tmp_path)
    monkeypatch.setattr(review, "EXPECTED_GPT_SHA256", review._sha256(paths["gpt"]))
    monkeypatch.setattr(review, "_is_git_ignored", lambda path, root: True)
    monkeypatch.setattr(review, "_git_head", lambda root: "a" * 40)
    output_dir = tmp_path / "preregistration_restricted" / "pilot_private_review_output"
    result = review.build_private_case_review(
        gpt_source=paths["gpt"],
        output_dir=output_dir,
        ready_path=paths["ready"],
        agreement_path=paths["agreement"],
        dimension_summary_path=paths["dimension"],
        raw_path=paths["raw"],
        fable_path=paths["fable"],
        fable_metadata_path=paths["metadata"],
        gpt_run_script_path=paths["run_script"],
        gpt_reference_path=paths["reference"],
        root=tmp_path,
    )
    return result, paths, output_dir


def test_canonical_gpt_source_is_the_cli_default():
    assert review._parser().parse_args([]).gpt_source == review.DEFAULT_GPT_SOURCE


def test_canonical_snapshot_has_1308_rows_and_retained_record(built):
    result, _, _ = built
    assert result["gpt"].row_count == 1308
    assert result["gpt"].unique_record_count == 1308
    assert result["snapshot"]["canonical_record_counts"] == {
        "2023/211": 1,
        "2023/211/a": 0,
        "2023/211/b": 0,
    }
    assert result["snapshot"]["pilot_records_unaffected"] is True
    assert result["reference_status"].startswith("Passed:")


def test_human_agreement_and_model_colours():
    unanimous = {coder: "same" for coder in review.CODERS}
    assert set(review._human_cell_colors(unanimous, "all_agree").values()) == {"green"}
    two = {"C01": "same", "C02": "same", "C03": "different"}
    assert review._human_cell_colors(two, "two_vs_one") == {
        "C01": "green", "C02": "green", "C03": "red"
    }
    split = {"C01": "one", "C02": "two", "C03": "three"}
    assert set(review._human_cell_colors(split, "split").values()) == {"amber"}
    humans = {coder: frozenset({value}) for coder, value in two.items()}
    assert review._model_cell_state(frozenset({"same"}), humans)[0] == "green"
    assert review._model_cell_state(frozenset({"other"}), humans)[0] == "red"
    split_humans = {coder: frozenset({value}) for coder, value in split.items()}
    assert review._model_cell_state(frozenset({"one"}), split_humans) == (
        "amber", "All three human sets differ; exact model match: C01."
    )


def test_comments_map_to_coders_and_blank_states_are_distinct(built):
    result, _, _ = built
    by_id = {row["record_id"]: row for row in result["rows"]}
    first = by_id["2021/103"]
    for coder in review.CODERS:
        assert f"synthetic note for {coder}" in first[f"{coder.lower()}_comments"]
    triggered_blank = by_id["SYN-001"]["c01_comments"]
    non_triggered = by_id["SYN-001"]["c02_comments"]
    assert "Blank response (field triggered)" in triggered_blank
    assert "Not triggered" in non_triggered
    assert result["comment_counts"] == {"C01": 1, "C02": 1, "C03": 1}


def test_stored_and_missing_rationale_handling():
    assert "verbatim rationale" in review._render_rationale(
        {"rationale": "verbatim rationale"}, "rationale"
    )
    assert review._render_rationale({}, None) == review.MISSING_RATIONALE
    assert review._render_rationale({"rationale": ""}, "rationale") == review.MISSING_RATIONALE


def test_workbook_shape_formatting_warning_and_annotations(built):
    result, _, output_dir = built
    workbook_path = output_dir / "pilot_private_case_review.xlsx"
    workbook = load_workbook(workbook_path)
    assert workbook.sheetnames == ["Case review", "Disagreement index", "Provenance and legend"]
    case = workbook["Case review"]
    assert (case.max_row, case.max_column) == (12, len(review.COLUMNS))
    assert case.freeze_panes == "D3"
    assert case.auto_filter.ref.endswith("12")
    assert all(cell.alignment.wrap_text for row in case.iter_rows(min_row=3) for cell in row)
    headers = {cell.value: cell.column for cell in case[2]}
    record_rows = {case.cell(row, 1).value: row for row in range(3, 13)}
    warning_cell = case.cell(record_rows["2021/103"], headers["C02 taxonomy issue type"])
    assert warning_cell.value == "None"
    assert warning_cell.fill.fgColor.rgb == review.COLORS["warning"]
    assert review.KNOWN_WARNING in warning_cell.comment.text
    split_model = case.cell(record_rows["SYN-002"], headers["Domain Fable"])
    assert split_model.fill.fgColor.rgb == review.COLORS["amber"]
    assert "exact model match: C01" in split_model.comment.text
    assert result["rows"][0]["domain_fable_rationale"].endswith("Stored overall rationale 0")
    provenance_values = [
        str(cell.value) for row in workbook["Provenance and legend"].iter_rows()
        for cell in row if cell.value
    ]
    assert any(review.CANONICAL_GPT_STATUS_TITLE in value for value in provenance_values)


def test_outputs_are_restricted_and_sources_unchanged(built):
    result, paths, output_dir = built
    assert len(result["rows"]) == 10
    assert len({row["record_id"] for row in result["rows"]}) == 10
    expected_keys = {column.key for column in review.COLUMNS}
    assert all(expected_keys <= set(row) for row in result["rows"])
    for path in result["outputs"]:
        assert path.resolve().is_relative_to(
            (paths["gpt"].parents[3] / "preregistration_restricted").resolve()
        )
        assert path.parent == output_dir
    legend = (output_dir / "pilot_private_case_review_legend.md").read_text(encoding="utf-8")
    assert review.CANONICAL_GPT_STATUS_TITLE in legend
    assert "preserved byte-for-byte" in legend
    assert "must be regenerated" not in legend
    for path, digest in result["source_hashes"].items():
        assert review._sha256(Path(path)) == digest
